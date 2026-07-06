from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .models import ActionType, Transaction, ValidationIssue
from .report_templates import ReportTemplate, load_report_templates


@dataclass(frozen=True, slots=True)
class BrokerLayout:
    broker: str
    required_fields: tuple[str, ...]
    field_aliases: dict[str, tuple[str, ...]]


@dataclass(frozen=True, slots=True)
class HeaderDetection:
    header_row_index: int
    broker: str
    column_map: dict[str, str]
    confidence: float
    template_name: str = ""


@dataclass(frozen=True, slots=True)
class HeaderPreview:
    source_file: str
    sheet: str
    header_row_index: int
    headers: list[str]
    sample_rows: list[list[str]]


BROKER_LAYOUTS = (
    BrokerLayout(
        broker="agis",
        required_fields=("trade_date", "action", "quantity", "net_amount"),
        field_aliases={
            "trade_date": ("Trade Date", "Execution Date", "Trade date", "Date"),
            "action": ("Transaction", "Action", "Activity"),
            "quantity": ("Quantity", "Qty", "Executed Quantity"),
            "price": ("Price ($)", "Price", "Trade Price"),
            "net_amount": ("Net Amount ($)", "Net Amount", "Net Proceeds"),
            "security_type": ("Security Type", "Type"),
            "settlement_date": ("Settlement Date", "Value Date"),
            "security_id": ("Cusip", "CUSIP", "Security ID", "ISIN"),
            "symbol": ("Security", "Symbol", "Ticker"),
            "security_name": ("Description", "Security Name", "Name"),
            "base_currency": ("Base Currency", "Currency", "Trade Currency"),
            "commission": ("Commissions ($)", "Commission", "Commissions"),
            "fees": ("Fees ($)", "Fees", "Other Fees"),
            "account_type": ("Account Type", "Account"),
        },
    ),
    BrokerLayout(
        broker="leumi",
        required_fields=("reference", "trade_date", "action", "quantity", "net_amount"),
        field_aliases={
            "reference": ("אסמכתא", "מספר אסמכתא", "אסמכתה"),
            "trade_date": ("תאריך ביצוע", "תאריך עסקה", "תאריך"),
            "action": ("פעולה", "סוג פעולה"),
            "security_id": ("מס' בורסה", "מספר בורסה", "מס' נייר", "מספר נייר"),
            "security_name": ('שם ני"ע', "שם נייר ערך", "שם נייר"),
            "quantity": ("כמות ביצוע", "כמות", "כמות נייר"),
            "price": ("שער ביצוע", "שער", "מחיר ביצוע"),
            "net_amount": ("תמורה נטו לפני מס", "תמורה נטו", "תמורה"),
            "currency": ("מטבע", "מטבע עסקה"),
            "commission": ("עמלות ודמי ניהול", "עמלות", "דמי ניהול"),
            "bank_reported_gain_loss": ("רווח/הפסד", "רווח הפסד", "רווח או הפסד"),
            "tax_rate": ("שעור המס", "שיעור המס"),
            "tax_withheld_local": ("מס שנוכה/הוחזר בארץ", "מס בארץ"),
            "tax_withheld_foreign": ('מס חו"ל בשקלים', "מס חול בשקלים"),
        },
    ),
)

GENERIC_REQUIRED_FIELDS = ("trade_date", "action", "quantity")
GENERIC_VALUE_FIELDS = ("net_amount", "price")
GENERIC_SECURITY_FIELDS = ("security_id", "symbol", "security_name")
GENERIC_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "reference": ("reference", "transaction id", "reference no", "doc number", "אסמכתא", "מספר אסמכתא", "אסמכתה"),
    "trade_date": (
        "trade date",
        "execution date",
        "date",
        "transaction date",
        "date executed",
        "activity date",
        "deal date",
        "booking date",
        "תאריך",
        "תאריך ביצוע",
        "תאריך עסקה",
        "תאריך מסחר",
    ),
    "settlement_date": ("settlement date", "settle date", "value date", "settlement", "תאריך ערך"),
    "action": (
        "action",
        "transaction",
        "transaction type",
        "activity",
        "activity type",
        "operation",
        "movement type",
        "side",
        "type",
        "פעולה",
        "תיאור פעולה",
        "סוג פעולה",
        "סוג",
    ),
    "quantity": (
        "quantity",
        "qty",
        "units",
        "shares",
        "share quantity",
        "executed quantity",
        "units shares",
        "כמות",
        "כמות נייר",
        "יחידות",
        "מניות",
    ),
    "price": (
        "price",
        "trade price",
        "execution price",
        "unit price",
        "average price",
        "execution rate",
        "שער",
        "מחיר",
        "שער ביצוע",
        "מחיר ביצוע",
    ),
    "net_amount": (
        "net amount",
        "net proceeds",
        "proceeds",
        "amount",
        "gross amount",
        "gross proceeds",
        "cash amount",
        "transaction amount",
        "deal amount",
        "תמורה נטו",
        "תמורה",
        "סכום",
        "שווי",
    ),
    "security_id": (
        "security id",
        "security number",
        "security no",
        "security code",
        "cusip",
        "isin",
        "מספר נייר",
        "מס נייר",
        "מספר בורסה",
        'מס" נייר',
    ),
    "symbol": ("symbol", "ticker", "ticker symbol", "security ticker", "symbol ticker", "code", "סימול", "סימול נייר", "קוד"),
    "security_name": (
        "security name",
        "security description",
        "description",
        "name",
        "instrument",
        "instrument name",
        "asset name",
        'שם ני"ע',
        "שם נייר ערך",
        "שם נייר",
        "תיאור",
    ),
    "currency": ("currency", "base currency", "trade currency", "currency code", "trade ccy", "ccy", "מטבע", "מטבע עסקה"),
    "report_currency": ("report currency", "statement currency", "currency report", "report ccy", "מטבע דיווח"),
    "commission": ("commission", "commissions", "broker fee", "transaction fee", "fee", "fees", "עמלה", "עמלות", "דמי ניהול"),
    "fees": ("other fees", "exchange fee", "fees", "charges", "חיובים", "דמי"),
    "bank_reported_gain_loss": ("gain/loss", "profit/loss", "realized pnl", "realized p&l", "reported gain", "רווח/הפסד", "רווח הפסד"),
}

GENERIC_FIELD_KEYWORDS: dict[str, tuple[tuple[str, ...], ...]] = {
    "reference": (("reference",), ("transaction", "id"), ("אסמכתא",)),
    "trade_date": (("trade", "date"), ("execution", "date"), ("date", "executed"), ("תאריך",), ("תאריך", "ביצוע")),
    "settlement_date": (("settlement", "date"), ("settle", "date"), ("value", "date"), ("תאריך", "ערך")),
    "action": (("transaction", "type"), ("activity", "type"), ("action",), ("operation",), ("side",), ("פעולה",), ("סוג", "פעולה")),
    "quantity": (("quantity",), ("qty",), ("units",), ("shares",), ("כמות",), ("יחידות",), ("מניות",)),
    "price": (("unit", "price"), ("trade", "price"), ("execution", "price"), ("price",), ("שער",), ("מחיר",)),
    "net_amount": (("net", "amount"), ("gross", "amount"), ("gross", "proceeds"), ("amount",), ("proceeds",), ("תמורה",), ("סכום",), ("שווי",)),
    "security_id": (("security", "id"), ("security", "number"), ("security", "code"), ("isin",), ("cusip",), ("מספר", "נייר"), ("מספר", "בורסה")),
    "symbol": (("ticker",), ("symbol",), ("ticker", "symbol"), ("סימול",), ("קוד",)),
    "security_name": (("security", "name"), ("instrument", "name"), ("description",), ("name",), ("שם", "נייר"), ("תיאור",)),
    "currency": (("trade", "currency"), ("currency",), ("currency", "code"), ("trade", "ccy"), ("ccy",), ("מטבע",)),
    "report_currency": (("report", "currency"), ("statement", "currency"), ("report", "ccy"), ("מטבע", "דיווח")),
    "commission": (("broker", "fee"), ("transaction", "fee"), ("commission",), ("עמלה",), ("עמלות",)),
    "fees": (("other", "fees"), ("exchange", "fee"), ("fees",), ("charges",), ("חיובים",)),
    "bank_reported_gain_loss": (("realized", "pnl"), ("realized", "p", "l"), ("gain", "loss"), ("profit", "loss"), ("רווח", "הפסד")),
}


def parse_workbooks(paths: list[str | Path]) -> tuple[list[Transaction], list[ValidationIssue]]:
    transactions: list[Transaction] = []
    issues: list[ValidationIssue] = []
    templates = load_report_templates()
    for path_like in paths:
        path = Path(path_like)
        parsed, file_issues = parse_workbook(path, templates=templates)
        transactions.extend(parsed)
        issues.extend(file_issues)
    return transactions, issues


def parse_workbook(path: Path, templates: list[ReportTemplate] | None = None) -> tuple[list[Transaction], list[ValidationIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    all_transactions: list[Transaction] = []
    issues: list[ValidationIssue] = []
    template_list = templates if templates is not None else load_report_templates()
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            detection = _detect_header(rows, template_list)
            if detection is None:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        message="Could not detect a supported report header",
                        source_file=path.name,
                        sheet=sheet.title,
                        row_number=1,
                        field="header",
                    )
                )
                continue

            headers = [_clean_header(v) for v in rows[detection.header_row_index - 1]]
            data_rows = rows[detection.header_row_index :]
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.dropna(how="all")

            if detection.broker == "agis":
                parsed, sheet_issues = _parse_agis(df, path.name, sheet.title, detection)
            elif detection.broker == "leumi":
                parsed, sheet_issues = _parse_leumi(df, path.name, sheet.title, detection)
            else:
                parsed, sheet_issues = _parse_generic(df, path.name, sheet.title, detection)
            all_transactions.extend(parsed)
            issues.extend(sheet_issues)
            if detection.broker == "generic":
                issues.append(
                    ValidationIssue(
                        severity="info",
                        message=f"Generic header mapping used (confidence {detection.confidence:.0%})",
                        source_file=path.name,
                        sheet=sheet.title,
                        row_number=detection.header_row_index,
                        field="header",
                        value=detection.template_name or "generic",
                    )
                )
    finally:
        workbook.close()

    return all_transactions, issues


def inspect_workbook_headers(path: Path) -> list[HeaderPreview]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    previews: list[HeaderPreview] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            best_index = _best_header_candidate_index(rows)
            if best_index is None:
                continue
            headers = [_clean_header(value) for value in rows[best_index - 1] if _clean_header(value)]
            sample_rows: list[list[str]] = []
            for row in rows[best_index : best_index + 3]:
                sample_rows.append([_text(cell) for cell in row[: min(len(row), 8)] if _text(cell)])
            previews.append(
                HeaderPreview(
                    source_file=path.name,
                    sheet=sheet.title,
                    header_row_index=best_index,
                    headers=headers,
                    sample_rows=sample_rows,
                )
            )
    finally:
        workbook.close()
    return previews


def _detect_header(rows: list[tuple[Any, ...]], templates: list[ReportTemplate]) -> HeaderDetection | None:
    for index, row in enumerate(rows, start=1):
        actual_headers = [_clean_header(v) for v in row if _clean_header(v)]
        normalized_map = {_normalize_header_text(header): header for header in actual_headers}
        for layout in BROKER_LAYOUTS:
            column_map = _match_layout(normalized_map, layout)
            if column_map is None:
                continue
            confidence = len(column_map) / len(layout.field_aliases)
            return HeaderDetection(index, layout.broker, column_map, confidence)
        template_detection = _match_template(normalized_map, templates)
        if template_detection is not None:
            return HeaderDetection(
                header_row_index=index,
                broker=template_detection.broker,
                column_map=template_detection.column_map,
                confidence=template_detection.confidence,
                template_name=template_detection.template_name,
            )
        generic_detection = _match_generic(normalized_map)
        if generic_detection is not None:
            return HeaderDetection(
                header_row_index=index,
                broker=generic_detection.broker,
                column_map=generic_detection.column_map,
                confidence=generic_detection.confidence,
                template_name=generic_detection.template_name,
            )
    return None


def _match_layout(normalized_map: dict[str, str], layout: BrokerLayout) -> dict[str, str] | None:
    column_map: dict[str, str] = {}
    for field_name, aliases in layout.field_aliases.items():
        matched_header = _find_matching_header(normalized_map, aliases)
        if matched_header:
            column_map[field_name] = matched_header
    if all(field_name in column_map for field_name in layout.required_fields):
        return column_map
    return None


def _find_matching_header(normalized_map: dict[str, str], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        actual = normalized_map.get(_normalize_header_text(alias))
        if actual:
            return actual
    return ""


def _match_template(normalized_map: dict[str, str], templates: list[ReportTemplate]) -> HeaderDetection | None:
    best_detection: HeaderDetection | None = None
    available_headers = set(normalized_map)
    for template in templates:
        matched_fields: dict[str, str] = {}
        for field_name, saved_header in template.field_map.items():
            actual = normalized_map.get(_normalize_header_text(saved_header))
            if actual:
                matched_fields[field_name] = actual
        if not _generic_fields_are_sufficient(matched_fields):
            continue
        score = len({field for field in matched_fields if matched_fields[field]}) / max(len(template.field_map), 1)
        if best_detection is None or score > best_detection.confidence:
            best_detection = HeaderDetection(
                header_row_index=0,
                broker=template.broker or "generic",
                column_map=matched_fields,
                confidence=score,
                template_name=template.name,
            )
    return best_detection


def _match_generic(normalized_map: dict[str, str]) -> HeaderDetection | None:
    field_map, field_scores = _collect_generic_matches(normalized_map)
    if not _generic_fields_are_sufficient(field_map):
        return None
    confidence = sum(field_scores.values()) / max(len(field_scores), 1)
    return HeaderDetection(0, "generic", field_map, confidence, template_name="")


def _collect_generic_matches(normalized_map: dict[str, str]) -> tuple[dict[str, str], dict[str, float]]:
    used_headers: set[str] = set()
    field_scores: dict[str, float] = {}
    field_map: dict[str, str] = {}
    field_order = [
        "trade_date",
        "action",
        "quantity",
        "symbol",
        "security_name",
        "security_id",
        "currency",
        "report_currency",
        "net_amount",
        "price",
        "commission",
        "fees",
        "reference",
        "bank_reported_gain_loss",
        "settlement_date",
    ]
    for field_name in field_order:
        aliases = GENERIC_FIELD_ALIASES[field_name]
        best_header = ""
        best_score = 0.0
        for normalized_header, actual_header in normalized_map.items():
            if normalized_header in used_headers:
                continue
            score = _generic_header_score(field_name, normalized_header, aliases)
            if score > best_score:
                best_header = actual_header
                best_score = score
        threshold = 0.7 if field_name in {"trade_date", "action", "quantity", "price", "net_amount"} else 0.66
        if best_header and best_score >= threshold:
            field_map[field_name] = best_header
            used_headers.add(_normalize_header_text(best_header))
            field_scores[field_name] = best_score
    return field_map, field_scores


def _generic_header_score(field_name: str, normalized_header: str, aliases: tuple[str, ...]) -> float:
    alias_score = max(_header_similarity(normalized_header, _normalize_header_text(alias)) for alias in aliases)
    header_tokens = _tokens(normalized_header)
    keyword_score = _keyword_match_score(header_tokens, GENERIC_FIELD_KEYWORDS.get(field_name, ()))
    return max(alias_score, keyword_score)


def _keyword_match_score(header_tokens: set[str], keyword_groups: tuple[tuple[str, ...], ...]) -> float:
    best_score = 0.0
    for group in keyword_groups:
        normalized_group = tuple(_normalize_header_text(token) for token in group if _normalize_header_text(token))
        if not normalized_group:
            continue
        group_tokens = set(normalized_group)
        overlap = len(header_tokens & group_tokens) / max(len(group_tokens), 1)
        if group_tokens.issubset(header_tokens):
            score = 0.86 if len(group_tokens) == 1 else 0.92
        else:
            score = overlap * (0.88 if len(group_tokens) == 1 else 0.9)
        best_score = max(best_score, score)
    return best_score


def _generic_fields_are_sufficient(field_map: dict[str, str]) -> bool:
    if not all(field in field_map for field in GENERIC_REQUIRED_FIELDS):
        return False
    if not any(field in field_map for field in GENERIC_VALUE_FIELDS):
        return False
    if not any(field in field_map for field in GENERIC_SECURITY_FIELDS):
        return False
    return True


def _parse_agis(df: pd.DataFrame, source_file: str, sheet: str, detection: HeaderDetection) -> tuple[list[Transaction], list[ValidationIssue]]:
    transactions: list[Transaction] = []
    issues: list[ValidationIssue] = []
    for position, row in df.iterrows():
        row_number = detection.header_row_index + 1 + int(position)
        action_raw = _text(_value(row, detection.column_map, "action"))
        trade_date = _parse_date(_value(row, detection.column_map, "trade_date"))
        quantity = _num(_value(row, detection.column_map, "quantity"))
        price_raw = _value(row, detection.column_map, "price")
        price = _num(price_raw)
        net_amount = _num(_value(row, detection.column_map, "net_amount"))
        security_type = _text(_value(row, detection.column_map, "security_type"))

        if not action_raw or not trade_date:
            continue
        if action_raw in {"Memo", "Charge", "Margin Int", "Journal"} and not quantity:
            continue
        if security_type and security_type.lower() != "equity":
            continue

        action_type = _map_agis_action(action_raw, quantity)
        transaction = Transaction(
            source_file=source_file,
            sheet=sheet,
            row_number=row_number,
            broker="Agis",
            trade_date=trade_date,
            settlement_date=_parse_date(_value(row, detection.column_map, "settlement_date")),
            action_raw=action_raw,
            action_type=action_type,
            security_id=_text(_value(row, detection.column_map, "security_id")),
            symbol=_text(_value(row, detection.column_map, "symbol")),
            security_name=_text(_value(row, detection.column_map, "security_name")),
            quantity=quantity,
            price=price,
            currency=_normalize_currency(_value(row, detection.column_map, "base_currency")),
            report_currency=_normalize_currency(_value(row, detection.column_map, "base_currency")),
            commission=_num(_value(row, detection.column_map, "commission")),
            fees=_num(_value(row, detection.column_map, "fees")),
            net_amount=net_amount,
            account_type=_text(_value(row, detection.column_map, "account_type")),
            description=_text(_value(row, detection.column_map, "security_name")),
            raw=_row_to_dict(row),
        )
        _validate_transaction(transaction, issues)
        if transaction.action_type in {ActionType.BUY, ActionType.SELL} and _is_missing(price_raw):
            issues.append(_issue(transaction, "error", "Missing price", "price", price_raw))
        if transaction.action_type != ActionType.IGNORE:
            transactions.append(transaction)
    return transactions, issues


def _parse_leumi(df: pd.DataFrame, source_file: str, sheet: str, detection: HeaderDetection) -> tuple[list[Transaction], list[ValidationIssue]]:
    transactions: list[Transaction] = []
    issues: list[ValidationIssue] = []
    for position, row in df.iterrows():
        row_number = detection.header_row_index + 1 + int(position)
        action_raw = _text(_value(row, detection.column_map, "action"))
        trade_date = _parse_date(_value(row, detection.column_map, "trade_date"), day_first=True)
        security_id = _text(_value(row, detection.column_map, "security_id"))
        security_name = _text(_value(row, detection.column_map, "security_name"))
        quantity = _num(_value(row, detection.column_map, "quantity"))
        price_raw = _value(row, detection.column_map, "price")
        price = _num(price_raw)
        net_amount = _num(_value(row, detection.column_map, "net_amount"))

        if not action_raw or not trade_date:
            continue
        if str(_value(row, detection.column_map, "reference", "")).startswith("סה"):
            continue

        action_type = _map_leumi_action(action_raw, quantity, net_amount)
        transaction = Transaction(
            source_file=source_file,
            sheet=sheet,
            row_number=row_number,
            broker="Leumi",
            trade_date=trade_date,
            action_raw=action_raw,
            action_type=action_type,
            security_id=security_id,
            symbol=security_id,
            security_name=security_name,
            quantity=quantity,
            price=price,
            currency=_normalize_currency(_value(row, detection.column_map, "currency")),
            report_currency="ILS",
            commission=_num(_value(row, detection.column_map, "commission")),
            fees=0.0,
            net_amount=net_amount,
            bank_reported_gain_loss=_optional_num(_value(row, detection.column_map, "bank_reported_gain_loss")),
            tax_rate=_optional_num(_value(row, detection.column_map, "tax_rate")),
            tax_withheld_local=_optional_num(_value(row, detection.column_map, "tax_withheld_local")),
            tax_withheld_foreign=_optional_num(_value(row, detection.column_map, "tax_withheld_foreign")),
            reference=_text(_value(row, detection.column_map, "reference")),
            description=security_name,
            raw=_row_to_dict(row),
        )
        _validate_transaction(transaction, issues)
        if transaction.action_type in {ActionType.BUY, ActionType.SELL} and _is_missing(price_raw):
            issues.append(_issue(transaction, "error", "Missing price", "price", price_raw))
        if transaction.action_type != ActionType.IGNORE:
            transactions.append(transaction)
    return transactions, issues


def _parse_generic(df: pd.DataFrame, source_file: str, sheet: str, detection: HeaderDetection) -> tuple[list[Transaction], list[ValidationIssue]]:
    transactions: list[Transaction] = []
    issues: list[ValidationIssue] = []
    for position, row in df.iterrows():
        row_number = detection.header_row_index + 1 + int(position)
        action_raw = _text(_value(row, detection.column_map, "action"))
        trade_date = _parse_flexible_date(_value(row, detection.column_map, "trade_date"))
        quantity_raw = _value(row, detection.column_map, "quantity")
        quantity = _num(quantity_raw)
        price_raw = _value(row, detection.column_map, "price")
        price = _num(price_raw)
        net_amount_raw = _value(row, detection.column_map, "net_amount")
        net_amount = _num(net_amount_raw)

        if not action_raw or not trade_date:
            continue
        if action_raw.startswith("סה") or _normalize_header_text(action_raw) in {"total", "summary"}:
            continue

        action_type = _map_generic_action(action_raw, quantity, net_amount)
        if action_type == ActionType.IGNORE:
            normalized_action = _normalize_header_text(action_raw)
            if any(token in normalized_action for token in ("consolidation", "איחוד")):
                action_type = ActionType.SPLIT_OUT
            elif any(token in normalized_action for token in ("stock split", "פיצול")):
                action_type = ActionType.SPLIT_IN
            elif any(token in normalized_action for token in ("dividend", "interest", "coupon", "דיבידנד", "ריבית")):
                action_type = ActionType.CASH
            elif quantity > 0 and net_amount < 0:
                action_type = ActionType.BUY
            elif quantity < 0 and net_amount > 0:
                action_type = ActionType.SELL
            elif quantity:
                action_type = ActionType.UNKNOWN
        currency = _normalize_currency(_value(row, detection.column_map, "currency"))
        report_currency = _normalize_currency(_value(row, detection.column_map, "report_currency"))
        if report_currency == "UNKNOWN":
            report_currency = currency
        transaction = Transaction(
            source_file=source_file,
            sheet=sheet,
            row_number=row_number,
            broker="Generic",
            trade_date=trade_date,
            settlement_date=_parse_flexible_date(_value(row, detection.column_map, "settlement_date")),
            action_raw=action_raw,
            action_type=action_type,
            security_id=_text(_value(row, detection.column_map, "security_id")),
            symbol=_text(_value(row, detection.column_map, "symbol")),
            security_name=_text(_value(row, detection.column_map, "security_name")),
            quantity=quantity,
            price=price,
            currency=currency,
            report_currency=report_currency,
            commission=_num(_value(row, detection.column_map, "commission")),
            fees=_num(_value(row, detection.column_map, "fees")),
            net_amount=net_amount,
            bank_reported_gain_loss=_optional_num(_value(row, detection.column_map, "bank_reported_gain_loss")),
            reference=_text(_value(row, detection.column_map, "reference")),
            description=_text(_value(row, detection.column_map, "security_name"))
            or _text(_value(row, detection.column_map, "symbol"))
            or _text(_value(row, detection.column_map, "security_id")),
            raw=_row_to_dict(row),
        )
        _validate_transaction(transaction, issues)
        if transaction.action_type in {ActionType.BUY, ActionType.SELL} and _is_missing(price_raw) and _is_missing(net_amount_raw):
            issues.append(_issue(transaction, "error", "Missing both price and net amount", "price", price_raw))
        if transaction.action_type != ActionType.IGNORE:
            transactions.append(transaction)
    return transactions, issues


def _map_agis_action(action: str, quantity: float) -> ActionType:
    normalized = action.strip().lower()
    if normalized == "purchase":
        return ActionType.BUY
    if normalized == "sale":
        return ActionType.SELL
    if normalized == "receive":
        return ActionType.TRANSFER_IN
    if normalized == "deliver":
        return ActionType.TRANSFER_OUT
    if "reverse" in normalized or "splt" in normalized or "split" in normalized:
        return ActionType.SPLIT_IN if quantity > 0 else ActionType.SPLIT_OUT
    if normalized == "stock movement":
        if quantity > 0:
            return ActionType.SPLIT_IN
        if quantity < 0:
            return ActionType.SPLIT_OUT
        return ActionType.CASH
    return ActionType.IGNORE


def _map_leumi_action(action: str, quantity: float, net_amount: float) -> ActionType:
    normalized = action.strip()
    if "קניה" in normalized or normalized == "הזמנה":
        return ActionType.BUY
    if "מכירה" in normalized or normalized in {"פדיון", "דמי ניכיון"}:
        return ActionType.SELL
    if "מקבל בהעברה" in normalized:
        return ActionType.TRANSFER_IN
    if normalized == "הקטנת הון":
        return ActionType.CAPITAL_REDUCTION
    if normalized == "פקיעה - נייר":
        return ActionType.EXPIRE
    if quantity == 0 and net_amount:
        return ActionType.CASH
    return ActionType.IGNORE


def _map_generic_action(action: str, quantity: float, net_amount: float) -> ActionType:
    normalized = _normalize_header_text(action)
    if any(token in normalized for token in ("buy", "purchase", "קניה", "קנייה", "acquire")):
        return ActionType.BUY
    if any(token in normalized for token in ("sell", "sale", "מכירה", "פדיון", "redeem")):
        return ActionType.SELL
    if any(token in normalized for token in ("transfer in", "receive", "מקבל בהעברה", "incoming")):
        return ActionType.TRANSFER_IN
    if any(token in normalized for token in ("transfer out", "deliver", "outgoing", "מסירה")):
        return ActionType.TRANSFER_OUT
    if "reverse split" in normalized or "split out" in normalized:
        return ActionType.SPLIT_OUT
    if "split in" in normalized or ("split" in normalized and quantity > 0):
        return ActionType.SPLIT_IN
    if "capital reduction" in normalized or "הקטנת הון" in normalized:
        return ActionType.CAPITAL_REDUCTION
    if "expire" in normalized or "פקיעה" in normalized:
        return ActionType.EXPIRE
    if quantity == 0 and net_amount:
        return ActionType.CASH
    return ActionType.IGNORE


def _validate_transaction(transaction: Transaction, issues: list[ValidationIssue]) -> None:
    if transaction.action_type in {ActionType.BUY, ActionType.SELL, ActionType.TRANSFER_IN}:
        if not transaction.quantity:
            issues.append(_issue(transaction, "error", "Missing or zero quantity", "quantity", transaction.quantity))
        if not transaction.security_id and not transaction.symbol and not transaction.security_name:
            issues.append(_issue(transaction, "error", "Missing security identifier", "security", ""))
    if transaction.action_type == ActionType.UNKNOWN:
        issues.append(_issue(transaction, "warning", f"Unknown action: {transaction.action_raw}", "action", transaction.action_raw))


def _issue(transaction: Transaction, severity: str, message: str, field: str, value: Any) -> ValidationIssue:
    return ValidationIssue(
        severity=severity,
        message=message,
        source_file=transaction.source_file,
        sheet=transaction.sheet,
        row_number=transaction.row_number,
        field=field,
        value=value,
    )


def _row_to_dict(row: pd.Series) -> dict[str, Any]:
    return {str(key): _serialize(value) for key, value in row.to_dict().items() if pd.notna(value)}


def _clean_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    for marker in ('"', "'", "׳", "״"):
        text = text.replace(marker, "")
    text = re.sub(r"[\(\)\[\]\{\}/\\|:;,\.\-_]+", " ", text)
    return " ".join(text.split())


def _tokens(value: Any) -> set[str]:
    return {token for token in _normalize_header_text(value).split() if token}


def _value(row: pd.Series, column_map: dict[str, str], field_name: str, default: Any = None) -> Any:
    header = column_map.get(field_name)
    if not header:
        return default
    return row.get(header, default)


def _text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _num(value: Any) -> float:
    if value is None or pd.isna(value) or value == "":
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").replace(" ", "")
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _optional_num(value: Any) -> float | None:
    if value is None or pd.isna(value) or value == "":
        return None
    return _num(value)


def _is_missing(value: Any) -> bool:
    return value is None or value == "" or pd.isna(value)


def _parse_date(value: Any, day_first: bool = False) -> datetime | None:
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, datetime):
        return value
    parsed = pd.to_datetime(value, dayfirst=day_first, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _parse_flexible_date(value: Any) -> datetime | None:
    parsed = _parse_date(value, day_first=False)
    if parsed is not None:
        return parsed
    return _parse_date(value, day_first=True)


def _normalize_currency(value: Any) -> str:
    text = _text(value)
    mapping = {
        "דולר": "USD",
        "ד.קנדי": "CAD",
        'ש"ח': "ILS",
        "שח": "ILS",
        "USD": "USD",
        "ILS": "ILS",
    }
    return mapping.get(text, text or "UNKNOWN")


def _header_similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.94
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = len(left_tokens & right_tokens) / max(len(left_tokens | right_tokens), 1)
    ratio = SequenceMatcher(None, left, right).ratio()
    return max(overlap, ratio)


def _best_header_candidate_index(rows: list[tuple[Any, ...]]) -> int | None:
    best_index: int | None = None
    best_score = -1.0
    for index, row in enumerate(rows[:25], start=1):
        headers = [_clean_header(value) for value in row if _clean_header(value)]
        normalized_map = {_normalize_header_text(header): header for header in headers}
        field_map, field_scores = _collect_generic_matches(normalized_map)
        recognized_fields = len(field_map)
        required_hits = sum(1 for field in GENERIC_REQUIRED_FIELDS if field in field_map)
        score = required_hits * 5 + recognized_fields * 2 + sum(field_scores.values()) + min(len(headers), 12) * 0.08
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _serialize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value
