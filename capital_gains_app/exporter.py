from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .dashboard import build_dashboard_summary
from .models import CalculationResult, CorporateActionRecord, Lot, RealizedMatch, Transaction, ValidationIssue


HEADER_FILL = PatternFill("solid", fgColor="1D2228")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DASHBOARD_FILL = PatternFill("solid", fgColor="050607")
CARD_FILL = PatternFill("solid", fgColor="111417")
ACCENT_FILL = PatternFill("solid", fgColor="2B3036")
THIN_BORDER = Border(bottom=Side(style="thin", color="545B64"))
CHART_BLUE = "8FD8FF"
CHART_PINK = "FF8FB8"
CHART_YELLOW = "FFE27A"
CHART_WHITE = "FFFFFF"
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00;"-"'
QTY_FORMAT = '#,##0.######;[Red]-#,##0.######;"-"'
DATE_FORMAT = "yyyy-mm-dd"


def export_result(result: CalculationResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"

    _write_dashboard(dashboard, result)
    summary = wb.create_sheet("Summary")
    _write_summary(summary, result)
    _write_realized(wb.create_sheet("Realized FIFO"), result.realized)
    _write_open_lots(wb.create_sheet("Open Positions"), result.open_lots)
    _write_transactions(wb.create_sheet("Transactions"), result.transactions)
    _write_corporate_actions(wb.create_sheet("Corporate Actions"), result.corporate_actions)
    _write_issues(wb.create_sheet("Validation Issues"), result.issues)

    for ws in wb.worksheets:
        _format_sheet(ws)

    wb.save(output_path)
    return output_path


def _write_dashboard(ws, result: CalculationResult) -> None:
    summary = build_dashboard_summary(result)
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=1, max_row=58, min_col=1, max_col=9):
        for cell in row:
            cell.fill = DASHBOARD_FILL
    ws["A1"] = "דשבורד רווחי הון"
    ws["A1"].font = Font(size=22, bold=True, color=CHART_WHITE)
    ws["A2"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = Font(color="A9B1B7")

    if result.exchange_rate:
        rate = result.exchange_rate
        ws["D1"] = "שער יציג דולר/שקל"
        ws["D1"].font = Font(bold=True, color=CHART_WHITE)
        ws["D2"] = rate.rate
        ws["D2"].number_format = "0.0000"
        ws["D2"].font = Font(size=18, bold=True, color=CHART_BLUE)
        ws["E2"] = (
            f"פורסם: {rate.published_date:%Y-%m-%d}; "
            f"תאריך מבוקש: {rate.requested_date:%Y-%m-%d}; "
            f"חודש אחורה: {rate.lookup_date:%Y-%m-%d}"
        )
        ws["E3"] = rate.note or rate.source
        ws["E2"].font = Font(color="A9B1B7")
        ws["E3"].font = Font(color="A9B1B7")

    insight_start = 10
    ws.cell(row=insight_start, column=1, value="5 תובנות מרכזיות")
    ws.cell(row=insight_start, column=1).font = Font(size=14, bold=True, color=CHART_WHITE)
    for index, insight in enumerate(summary.key_insights, start=1):
        cell = ws.cell(row=insight_start + index, column=1, value=f"{index}. {insight}")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")
        cell.font = Font(color=CHART_WHITE)
        for col in range(1, 5):
            ws.cell(row=insight_start + index, column=col).fill = CARD_FILL

    kpis = [
        ("תנועות", summary.total_transactions),
        ("ניירות", summary.unique_securities),
        ("שורות FIFO", summary.realized_rows),
        ("פוזיציות פתוחות", summary.open_lots),
        ("התראות", summary.issue_count),
        ("שורות מוסקות", summary.inferred_rows),
        ("אירועי הון", summary.corporate_actions),
    ]
    for index, (label, value) in enumerate(kpis):
        row = 5 + (index // 3) * 3
        col = 1 + (index % 3) * 3
        ws.cell(row=row, column=col, value=label)
        ws.cell(row=row + 1, column=col, value=value)
        for offset in range(3):
            cell = ws.cell(row=row, column=col + offset)
            cell.fill = ACCENT_FILL
            cell.border = THIN_BORDER
            cell.font = Font(bold=True, color=CHART_WHITE)
            ws.cell(row=row + 1, column=col + offset).fill = CARD_FILL
        ws.cell(row=row + 1, column=col).font = Font(size=18, bold=True, color=CHART_BLUE)

    gain_start = 18
    ws.cell(row=gain_start, column=1, value="רווח/הפסד לפי מטבע")
    ws.cell(row=gain_start, column=1).font = Font(size=14, bold=True, color=CHART_WHITE)
    ws.append([])
    ws.cell(row=gain_start + 1, column=1, value="מטבע")
    ws.cell(row=gain_start + 1, column=2, value="רווח/הפסד")
    for row_index, (currency, value) in enumerate(summary.gain_by_currency, start=gain_start + 2):
        ws.cell(row=row_index, column=1, value=currency)
        ws.cell(row=row_index, column=2, value=value)
        ws.cell(row=row_index, column=1).font = Font(color=CHART_WHITE)
        ws.cell(row=row_index, column=2).font = Font(color=CHART_BLUE if value >= 0 else CHART_PINK)

    if summary.gain_by_currency:
        chart = BarChart()
        chart.title = "רווח/הפסד לפי מטבע"
        chart.y_axis.title = "רווח/הפסד"
        chart.x_axis.title = "מטבע"
        data = Reference(ws, min_col=2, min_row=gain_start + 1, max_row=gain_start + 1 + len(summary.gain_by_currency))
        cats = Reference(ws, min_col=1, min_row=gain_start + 2, max_row=gain_start + 1 + len(summary.gain_by_currency))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _style_chart_series(chart, [CHART_BLUE])
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "D12")

    top_start = 28
    ws.cell(row=top_start, column=1, value="ניירות בולטים לפי רווח/הפסד מוחלט")
    ws.cell(row=top_start, column=1).font = Font(size=14, bold=True, color=CHART_WHITE)
    ws.cell(row=top_start + 1, column=1, value="נייר")
    ws.cell(row=top_start + 1, column=2, value="מטבע")
    ws.cell(row=top_start + 1, column=3, value="רווח/הפסד")
    for row_index, (label, currency, value) in enumerate(summary.top_securities, start=top_start + 2):
        ws.cell(row=row_index, column=1, value=label[:35])
        ws.cell(row=row_index, column=2, value=currency)
        ws.cell(row=row_index, column=3, value=value)
        ws.cell(row=row_index, column=1).font = Font(color=CHART_WHITE)
        ws.cell(row=row_index, column=2).font = Font(color=CHART_WHITE)
        ws.cell(row=row_index, column=3).font = Font(color=CHART_BLUE if value >= 0 else CHART_PINK)
    if summary.top_securities:
        chart = BarChart()
        chart.title = "ניירות בולטים"
        chart.y_axis.title = "רווח/הפסד"
        data = Reference(ws, min_col=3, min_row=top_start + 1, max_row=top_start + 1 + len(summary.top_securities))
        cats = Reference(ws, min_col=1, min_row=top_start + 2, max_row=top_start + 1 + len(summary.top_securities))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _style_chart_series(chart, [CHART_PINK])
        chart.height = 8
        chart.width = 15
        ws.add_chart(chart, "E22")

    action_start = 42
    ws.cell(row=action_start, column=1, value="פילוח פעולות")
    ws.cell(row=action_start, column=1).font = Font(size=14, bold=True, color=CHART_WHITE)
    ws.cell(row=action_start + 1, column=1, value="פעולה")
    ws.cell(row=action_start + 1, column=2, value="כמות")
    for row_index, (label, value) in enumerate(summary.action_counts, start=action_start + 2):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
        ws.cell(row=row_index, column=1).font = Font(color=CHART_WHITE)
        ws.cell(row=row_index, column=2).font = Font(color=CHART_YELLOW)
    if summary.action_counts:
        chart = PieChart()
        chart.title = "פעולות"
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        data = Reference(ws, min_col=2, min_row=action_start + 1, max_row=action_start + 1 + len(summary.action_counts))
        cats = Reference(ws, min_col=1, min_row=action_start + 2, max_row=action_start + 1 + len(summary.action_counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _style_chart_series(chart, [CHART_WHITE, CHART_BLUE, CHART_PINK, CHART_YELLOW])
        chart.height = 8
        chart.width = 10
        ws.add_chart(chart, "D36")

    for row in [gain_start + 1, top_start + 1, action_start + 1]:
        for cell in ws[row]:
            if cell.value:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")


def _write_summary(ws, result: CalculationResult) -> None:
    ws.append(["מדד", "ערך"])
    totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in result.realized:
        totals[row.currency]["Proceeds"] += row.proceeds
        totals[row.currency]["Cost Basis"] += row.cost_basis
        totals[row.currency]["Gain/Loss"] += row.gain_loss
        if row.bank_reported_gain_loss is not None:
            totals[row.currency]["Bank Reported Gain/Loss"] += row.bank_reported_gain_loss

    ws.append(["תנועות", len(result.transactions)])
    ws.append(["ניירות ייחודיים", build_dashboard_summary(result).unique_securities])
    ws.append(["שורות FIFO ממומשות", len(result.realized)])
    ws.append(["פוזיציות פתוחות", len(result.open_lots)])
    ws.append(["אירועי הון", len(result.corporate_actions)])
    ws.append(["התראות תקינות", len(result.issues)])
    for index, insight in enumerate(build_dashboard_summary(result).key_insights, start=1):
        ws.append([f"תובנה {index}", insight])
    if result.exchange_rate:
        rate = result.exchange_rate
        ws.append(["תאריך מבוקש לשער דולר", rate.requested_date])
        ws.append(["תאריך יעד חודש אחורה", rate.lookup_date])
        ws.append(["תאריך פרסום שער", rate.published_date])
        ws.append(["שער יציג דולר/שקל", rate.rate])
        ws.append(["מקור שער", rate.source])
        ws.append(["הערת שער", rate.note])
    ws.append([])
    ws.append(["מטבע", "תמורה", "עלות", "רווח/הפסד", "רווח/הפסד מדווח בנק", "פער"])
    for currency, values in sorted(totals.items()):
        bank_gain = values.get("Bank Reported Gain/Loss", 0.0)
        gain = values.get("Gain/Loss", 0.0)
        ws.append(
            [
                currency,
                values.get("Proceeds", 0.0),
                values.get("Cost Basis", 0.0),
                gain,
                bank_gain if bank_gain else None,
                gain - bank_gain if bank_gain else None,
            ]
        )


def _write_realized(ws, rows: Iterable[RealizedMatch]) -> None:
    ws.append(
        [
            "sale_date",
            "security_key",
            "security_id",
            "symbol",
            "security_name",
            "quantity",
            "proceeds",
            "cost_basis",
            "gain_loss",
            "currency",
            "buy_date",
            "buy_source_file",
            "buy_row",
            "sale_source_file",
            "sale_row",
            "inferred",
            "action_raw",
            "bank_reported_gain_loss",
            "difference_vs_bank",
        ]
    )
    for row in rows:
        bank = row.bank_reported_gain_loss
        ws.append(
            [
                row.sale_date,
                row.security_key,
                row.security_id,
                row.symbol,
                row.security_name,
                row.quantity,
                row.proceeds,
                row.cost_basis,
                row.gain_loss,
                row.currency,
                row.buy_date,
                row.buy_source_file,
                row.buy_row,
                row.sale_source_file,
                row.sale_row,
                row.inferred,
                row.action_raw,
                bank,
                row.gain_loss - bank if bank is not None else None,
            ]
        )


def _write_open_lots(ws, rows: Iterable[Lot]) -> None:
    ws.append(
        [
            "security_key",
            "security_id",
            "symbol",
            "security_name",
            "acquired_date",
            "quantity",
            "unit_cost",
            "total_cost",
            "currency",
            "source_file",
            "source_row",
            "inferred",
            "notes",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.security_key,
                row.security_id,
                row.symbol,
                row.security_name,
                row.acquired_date,
                row.quantity,
                row.unit_cost,
                row.total_cost,
                row.currency,
                row.source_file,
                row.source_row,
                row.inferred,
                row.notes,
            ]
        )


def _write_transactions(ws, rows: Iterable[Transaction]) -> None:
    ws.append(
        [
            "trade_date",
            "broker",
            "source_file",
            "sheet",
            "row_number",
            "action_raw",
            "action_type",
            "security_key",
            "security_id",
            "symbol",
            "security_name",
            "quantity",
            "price",
            "currency",
            "report_currency",
            "net_amount",
            "commission",
            "fees",
            "bank_reported_gain_loss",
            "reference",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.trade_date,
                row.broker,
                row.source_file,
                row.sheet,
                row.row_number,
                row.action_raw,
                row.action_type.value,
                row.inventory_key,
                row.security_id,
                row.symbol,
                row.security_name,
                row.quantity,
                row.price,
                row.currency,
                row.report_currency,
                row.net_amount,
                row.commission,
                row.fees,
                row.bank_reported_gain_loss,
                row.reference,
            ]
        )


def _write_corporate_actions(ws, rows: Iterable[CorporateActionRecord]) -> None:
    ws.append(
        [
            "action_date",
            "action_type",
            "old_key",
            "new_key",
            "old_quantity",
            "new_quantity",
            "ratio",
            "source_file",
            "row_numbers",
            "notes",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.action_date,
                row.action_type,
                row.old_key,
                row.new_key,
                row.old_quantity,
                row.new_quantity,
                row.ratio,
                row.source_file,
                row.row_numbers,
                row.notes,
            ]
        )


def _write_issues(ws, rows: Iterable[ValidationIssue]) -> None:
    ws.append(["severity", "message", "source_file", "sheet", "row_number", "field", "value"])
    for row in rows:
        ws.append([row.severity, row.message, row.source_file, row.sheet, row.row_number, row.field, row.value])


def _style_chart_series(chart, colors: list[str]) -> None:
    if not colors:
        return
    for index, series in enumerate(chart.series):
        color = colors[index % len(colors)]
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color
        if isinstance(chart, PieChart):
            series.dPt = []
            for point_index in range(32):
                point = DataPoint(idx=point_index)
                point.graphicalProperties.solidFill = colors[point_index % len(colors)]
                series.dPt.append(point)


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = DATE_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(cell.value, float):
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if "quantity" in header or "ratio" in header or "price" in header or "unit" in header:
                    cell.number_format = QTY_FORMAT
                else:
                    cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.value is not None:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=cell.alignment.wrap_text)

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 10), 42)
    if ws.title != "Dashboard":
        ws.auto_filter.ref = ws.dimensions
