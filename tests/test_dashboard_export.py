from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from capital_gains_app.dashboard import build_dashboard_summary
from capital_gains_app.exchange_rates import one_month_back, parse_boi_exchange_rate_csv
from capital_gains_app.exporter import export_result
from capital_gains_app.fifo import calculate_fifo
from capital_gains_app.models import ActionType, ExchangeRateSnapshot, Transaction


def tx(row, trade_date, action, qty, price, net):
    return Transaction(
        source_file="single_security.xlsx",
        sheet="Sheet1",
        row_number=row,
        broker="Test",
        trade_date=datetime.fromisoformat(trade_date),
        action_raw=action.value,
        action_type=action,
        security_id="AAA",
        symbol="AAA",
        security_name="Single Security",
        quantity=qty,
        price=price,
        currency="USD",
        report_currency="USD",
        net_amount=net,
    )


class DashboardExportTests(unittest.TestCase):
    def test_dashboard_and_export_support_single_security_report(self):
        result = calculate_fifo(
            [
                tx(1, "2024-01-01", ActionType.BUY, 100, 10, -1000),
                tx(2, "2024-02-01", ActionType.SELL, -40, 15, 600),
            ]
        )
        result.exchange_rate = ExchangeRateSnapshot(
            requested_date=date(2024, 2, 15),
            lookup_date=date(2024, 1, 15),
            published_date=date(2024, 1, 15),
            currency_pair="USD/ILS",
            rate=3.72,
            source="test",
        )

        summary = build_dashboard_summary(result)
        self.assertEqual(summary.unique_securities, 1)
        self.assertEqual(summary.total_transactions, 2)
        self.assertEqual(summary.realized_rows, 1)
        self.assertEqual(len(summary.key_insights), 5)
        self.assertIn("נייר ערך אחד", summary.key_insights[-1])

        with TemporaryDirectory() as tmp:
            output = Path(tmp) / "single_security_report.xlsx"
            export_result(result, output)
            workbook = load_workbook(output, data_only=True)
            self.assertIn("Dashboard", workbook.sheetnames)
            self.assertIn("Realized FIFO", workbook.sheetnames)
            self.assertIn("Validation Issues", workbook.sheetnames)
            self.assertEqual(workbook["Dashboard"]["A1"].value, "דשבורד רווחי הון")
            self.assertTrue(workbook["Dashboard"].sheet_view.rightToLeft)
            self.assertEqual(workbook.properties.creator, "Capital Gains")

    def test_bank_of_israel_csv_parser_and_one_month_back(self):
        raw = (
            "SERIES_CODE,TIME_PERIOD,OBS_VALUE\n"
            "RER_USD_ILS,2024-01-14,3.70\n"
            "RER_USD_ILS,2024-01-15,3.72\n"
        )
        rows = parse_boi_exchange_rate_csv(raw)
        self.assertEqual(rows[-1], (date(2024, 1, 15), 3.72))
        self.assertEqual(one_month_back(date(2024, 3, 31)), date(2024, 2, 29))

    def test_export_includes_validation_sheet_even_without_issues(self):
        result = calculate_fifo([tx(1, "2024-01-01", ActionType.BUY, 10, 10, -100)])
        result.issues = []

        with TemporaryDirectory() as tmp:
            output = Path(tmp) / "validation_sheet.xlsx"
            export_result(result, output)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Validation Issues"]
            self.assertEqual(sheet["A1"].value, "חומרה")
            self.assertEqual(sheet["B2"].value, "לא נמצאו התראות או חריגים.")


if __name__ == "__main__":
    unittest.main()
