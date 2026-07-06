from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from capital_gains_app.parsers import parse_workbook


class ParserHeaderTests(unittest.TestCase):
    def test_parses_agis_style_report_with_header_aliases(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "alias_report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["Intro", "", "", "", "", "", "", "", ""])
            sheet.append(
                [
                    "Execution Date",
                    "Action",
                    "Qty",
                    "Net Amount",
                    "Trade Price",
                    "Security ID",
                    "Symbol",
                    "Security Name",
                    "Currency",
                ]
            )
            sheet.append(["2024-01-03", "Sale", -10, 150.0, 15.0, "AAA111", "AAA", "Alpha Asset", "USD"])
            workbook.save(workbook_path)

            transactions, issues = parse_workbook(workbook_path)

        self.assertEqual(len(transactions), 1)
        self.assertEqual(transactions[0].broker, "Agis")
        self.assertEqual(transactions[0].symbol, "AAA")
        self.assertEqual(transactions[0].security_id, "AAA111")
        self.assertAlmostEqual(transactions[0].price, 15.0)
        self.assertFalse([issue for issue in issues if issue.severity == "error"])

    def test_parses_leumi_style_report_with_header_aliases(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "leumi_alias_report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "תנועות"
            sheet.append(["כותרת", "", "", "", "", "", ""])
            sheet.append(["מספר אסמכתא", "תאריך עסקה", "סוג פעולה", "מספר נייר", "כמות", "מחיר ביצוע", "תמורה נטו"])
            sheet.append(["12345", "06/01/2024", "קניה", "1087824", 5, 120.5, -602.5])
            workbook.save(workbook_path)

            transactions, issues = parse_workbook(workbook_path)

        self.assertEqual(len(transactions), 1)
        self.assertEqual(transactions[0].broker, "Leumi")
        self.assertEqual(transactions[0].security_id, "1087824")
        self.assertAlmostEqual(transactions[0].price, 120.5)
        self.assertFalse([issue for issue in issues if issue.severity == "error"])

    def test_parses_generic_report_via_flexible_header_matching(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "generic_report.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            sheet.append(["Cover", "", "", "", "", ""])
            sheet.append(["Trade Date", "Transaction Type", "Ticker", "Units", "Unit Price", "Amount"])
            sheet.append(["2024-03-01", "Buy", "MSFT", 3, 200.0, -600.0])
            workbook.save(workbook_path)

            transactions, issues = parse_workbook(workbook_path)

        self.assertEqual(len(transactions), 1)
        self.assertEqual(transactions[0].broker, "Generic")
        self.assertEqual(transactions[0].symbol, "MSFT")
        self.assertAlmostEqual(transactions[0].price, 200.0)
        self.assertFalse([issue for issue in issues if issue.severity == "error"])

    def test_generic_header_detection_prefers_recognizable_header_row(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "generic_header_priority.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            sheet.append(["Summary", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight"])
            sheet.append(["Date Executed", "Activity Type", "Security / Ticker", "Units / Shares", "Gross Proceeds", "Trade CCY"])
            sheet.append(["2024-03-01", "Sell", "MSFT", -3, 650.0, "USD"])
            workbook.save(workbook_path)

            transactions, issues = parse_workbook(workbook_path)

        self.assertEqual(len(transactions), 1)
        self.assertEqual(transactions[0].broker, "Generic")
        self.assertEqual(transactions[0].symbol, "MSFT")
        self.assertEqual(transactions[0].currency, "USD")
        self.assertFalse([issue for issue in issues if issue.severity == "error"])

    def test_generic_parser_infers_buy_sell_from_quantity_and_amount_signs(self) -> None:
        with TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "generic_inferred_action.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Activity"
            sheet.append(["Activity Date", "Movement Type", "Ticker Symbol", "Share Quantity", "Transaction Amount"])
            sheet.append(["2024-03-01", "Booked", "AAPL", 4, -720.0])
            sheet.append(["2024-03-02", "Booked", "AAPL", -2, 390.0])
            workbook.save(workbook_path)

            transactions, issues = parse_workbook(workbook_path)

        self.assertEqual(len(transactions), 2)
        self.assertEqual(transactions[0].action_type.value, "BUY")
        self.assertEqual(transactions[1].action_type.value, "SELL")
        self.assertFalse([issue for issue in issues if issue.severity == "error"])


if __name__ == "__main__":
    unittest.main()
